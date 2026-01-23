from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Any
from typing import Tuple
from col_descriptors import get_cargo_all_col_descriptions


def _open_or_create_workbook(filepath: str) -> Workbook:
    """
    Checks if a workbook is already available or creates one based the provided
    `filepath`. The method `save(filepath)` still needs to be used on the returned
    object to save it to the disk.
    
    :param filepath: The filepath of the workbook to create or open.
    :type filepath: str
    :return: Either the newly created Workbook object or the opened Workbook object.
    :rtype: Workbook
    """
    path = Path(filepath)

    if path.exists():
        return load_workbook(path)
    else:
        return Workbook()

def _reset_ws_contents(ws):
    """
    Delete the values of the cells of a Worksheet and reset the style.
    
    :param ws: The Worksheet to work with.
    """
    # delete existing contents
    for row in ws.rows:
        for cell in row:
            # Resetting the cell.
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.fill = PatternFill()
            cell.protection = Protection()
            cell.number_format = "General"

def _style_header(ws):
    for col in ["A", "B"]:
        ws[f"{col}1"].font = Font(
            size = 14,
            bold = True,
        )

        ws[f"{col}1"].alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        thin = Side(style='thin')
        ws[f"{col}1"].border = Border(
            top=thin,
            bottom = Side(style='medium'),
            left=thin,
            right=thin
        )

        ws[f"{col}1"].fill = PatternFill(
            fill_type="solid",
            start_color="d84e85",
            end_color="d84e85"
        )

def _write_header(ws, headers: list[str]):
    ws["A1"] = headers[0]
    ws["B1"] = headers[1]

def _write_vulnerability_headers(ws, headers: list[str]):
    col = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
        "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
        "U", "V", "W", "X", "Y", "Z",
    ]

    for header_index in range(0, len(headers)):
        ws[f"{col[header_index]}1"] = headers[header_index][0]

def _style_vulnerability_headers(ws):
    '''font, size, fill, alignment'''
    for col in ws.columns:
        for cell in col:
            if cell.row == 2:
                break
        
            cell.font = Font(
                size = 14,
                bold = True,
            )

            cell.alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

            thin = Side(style='thin')
            cell.border = Border(
                top=thin,
                bottom = Side(style='medium'),
                left=thin,
                right=thin
            )

            cell.fill = PatternFill(
                fill_type="solid",
                start_color="d84e85",
                end_color="d84e85"
            )

def _write_col_descriptors(wb: Workbook, descriptors: list[Tuple[str, str]]) -> None:
    """
    Write column descriptors in the provided `wb` object, in a sheet named `Cargo Column Descriptors`.
    If the sheet already exists, delete its contents first.
    
    :param wb: The `Workbook` object to write to.
    :type wb: Workbook
    :param descriptors: The list of column descriptors to write.
    :type descriptors: list[Tuple[str, str]]
    """
    def __write_to_sheet__(ws):        
        for index in range(0, len(descriptors)):
            titleLoc = f"A{index+2}"
            descLoc = f"B{index+2}"

            ws[titleLoc] = descriptors[index][0]
            ws[descLoc] = descriptors[index][1]
    
    def __style__title_column__(ws):
        for column in ws.columns:
            for cell in column:
                if cell.coordinate == "A1":
                    continue
                if not cell.value:
                    break
                if cell.col_idx == 2:
                    break
                
                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="e8650d",
                    end_color="e8650d"
                )
            break
             
    sheetTitle = "Cargo Row Descriptors"
    if sheetTitle in wb.sheetnames:
        ws = wb[sheetTitle]
        _reset_ws_contents(ws)
        _write_header(ws, ["Row Name", "Row Description"])
        __write_to_sheet__(ws)
        _style_header(ws)
        __style__title_column__(ws)

    else:
        ws = wb.create_sheet(title = sheetTitle)
        _write_header(ws, ["Row Name", "Row Description"])
        __write_to_sheet__(ws)
        _style_header(ws)
        __style__title_column__(ws)

def _workbook_cleanup_(wb: Workbook):
    """
    Remove the default sheet in the workbook named `Sheet`, if exists.
    
    :param wb: The `Workbook` object to manipulate.
    :type wb: Workbook
    """
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
def _write_cargo_vulnerability_results_(cargo_results: dict[str, Any], wb: Workbook):
    column_letters = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
        "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
        "U", "V", "W", "X", "Y", "Z",
    ]

    sheetTitle = "Cargo Audit Results"
    if sheetTitle in wb.sheetnames:
        ws = wb[sheetTitle]
        _reset_ws_contents(ws)
    else:
        ws = wb.create_sheet(sheetTitle)
    
    if "vulnerabilities" not in cargo_results:
        ws["A1"] = "No vulnerabilities found"
        return
    
    _write_header(ws, ["Fields", "Details"])
    vulns = cargo_results["vulnerabilities"]
    col_descriptions = get_cargo_all_col_descriptions()
    _write_vulnerability_headers(ws, col_descriptions)
    _style_vulnerability_headers(ws)
    
    # Get the indexes to write to
    # Write titles, then values, then merge and blacken
    for vuln_index in range(0, len(vulns)):
        row = vuln_index + 2
        vuln_values = list(vulns[vuln_index].values())
        for value_index in range(0, len(vuln_values)):
            ws[f"{column_letters[value_index]}{row}"] = vuln_values[value_index]

def _resize_columns_(wb: Workbook):
    for name in wb.sheetnames:
        ws = wb[name]
        
        for cells in ws.columns:
            max_length = 0
            column_letter = cells[0].column_letter

            for cell in cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length


def write_to_workbook(cargo_results: dict[str, Any], 
        filepath: str, 
        col_descriptors: bool = True) -> None:
    """
    Write the cargo results to a XLSX file. The method optionally writes the
    column descriptors to the file in a new tab.
    
    :param cargo_results: The processed results obtained from `cargo-audit`.
    :type cargo_results: dict[str, Any]
    :param filepath: The filepath of the file to write to.
    :type filepath: str
    :param col_descriptors: Whether or not to write the column descriptors tab.
    :type col_descriptors: bool. Defaults to True.
    """
    activeWb = _open_or_create_workbook(filepath)

    if col_descriptors:
        _write_col_descriptors(activeWb, get_cargo_all_col_descriptions())
        
    _write_cargo_vulnerability_results_(cargo_results, activeWb)
    _workbook_cleanup_(activeWb)
    _resize_columns_(activeWb)
    activeWb.save(filepath)


